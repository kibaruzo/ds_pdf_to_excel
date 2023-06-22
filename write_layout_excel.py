import fitz
from openpyxl import Workbook
import os
from config import output_path, pdf_path, entered_start_page, entered_end_page

def get_page_layout(pdf_path, start_page, end_page):
    doc = fitz.open(pdf_path)

    named_layout = []
    for page_number in range(start_page, end_page + 1):
        page = doc[page_number]
        layout = page.get_textpage().extractBLOCKS()

        for element in layout:
            element_dict = {
                'page_number': page_number - start_page + 1,
                'x0': element[0],
                'y0': element[1],
                'x1': element[2],
                'y1': element[3],
                'text': element[4],
                #'block_no': element[5],
                #'block_type': element[6],
            }

            named_layout.append(element_dict)

    return named_layout

def proper_case(text):
    words = text.split()
    proper_words = [word.capitalize() for word in words]
    return ' '.join(proper_words)

def write_data_to_excel(layout_elements):
    
    # Create a new workbook
    workbook = Workbook()
    worksheet = workbook.active
    
    for page_number in sorted(set(element['page_number'] for element in layout_elements)):
        
        # Filter layout elements for the current page
        page_elements = [element for element in layout_elements if element['page_number'] == page_number]
    
        # Separate elements based on x0 value
        first_sort = [element for element in page_elements if element['x0'] <= 340]
        second_sort = [element for element in page_elements if element['x0'] > 340]

        # Sort the first set of elements by y0
        first_sort = sorted(first_sort, key=lambda e: e['y0'])

        # Sort the second set of elements by y0
        second_sort = sorted(second_sort, key=lambda e: e['y0'])

        # Concatenate the sorted elements
        sorted_elements = first_sort + second_sort
        
        # Find the index of the element with "KEYWORDS:" in its text
        keywords_index = next((i for i, e in enumerate(sorted_elements) if e['text'].startswith('KEYWORDS:')), None)

        # Move the element with "KEYWORDS:" to the end of the list
        if keywords_index is not None:
            sorted_elements.append(sorted_elements.pop(keywords_index))

        #IF page 1 rename sheet1 to unit name
        if sorted_elements[0]['page_number'] == 1:
            #odd page number
            worksheet.title = proper_case(sorted_elements[0]['text']) 
        
        #If odd page rename active sheet with unit name
        elif sorted_elements[0]['page_number'] % 2 != 0:
            worksheet = workbook.create_sheet(title=proper_case(sorted_elements[0]['text']))
           
        for element in sorted_elements:
            print(element)
        
        # Set ABILITIES flag to false
        flag = False       
        # Iterate over the sorted layout elements
        for element in sorted_elements:
            text = element['text']
            lines = [text]

            if flag == False or element['text'].startswith('INVULNERABLE SAVE'):
                # Split the text into lines
                lines = text.splitlines()
            
            if element['text'].startswith('ABILITIES') or element['page_number'] % 2 == 0:
                flag = True
                        
            
            # Write each line to a separate cell
            row = worksheet.max_row + 1
            for i, line in enumerate(lines):
                cell = worksheet.cell(row=row, column=i + 1)
                cell.value = line
        
    # Save the workbook with the PDF file name
    
    workbook.save(output_path)

#  Usage #########################################################################################################

start_page = entered_start_page - 1  # Start page number (0-based index)
end_page = entered_end_page -1  # End page number (0-based index)
layout_elements = get_page_layout(pdf_path, start_page, end_page)

write_data_to_excel(layout_elements)